import openpyxl, os

wb = openpyxl.load_workbook('updatedProduceSales.xlsx')   # Open your excel file which has the wrong values
sheet = wb['Sheet']

t = 1
while sheet.cell(row=t, column=1).value != None:          # calculate the amount of rows
    t = t+1
print(t)
New_price = {'Celery':'1,19','Garlic':'3,07','Lemon':'1,27'}       # keys are names whose values changes

for i in range(1, t):

    productname = sheet.cell(row=i, column=1).value     # loop through all rows, if the val of row t col 1 == key in dict
                                                        # than change the value of row t col 2 with the value of the associated key in dict
    if productname in New_price:
        sheet.cell(row=i, column=2).value = New_price[productname]
wb.save('Juistefiles.xlsx')                             # save program to different file 

"""Written by Mohamed Taouil""""